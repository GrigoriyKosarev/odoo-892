import base64
import io
from datetime import datetime
from odoo import fields, models, _, api
from odoo.exceptions import UserError

# Try to import Excel parsing libraries
try:
    import xlrd
    from xlrd import xldate
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


class MrpProductionSheduleImportWizard(models.TransientModel):
    _name = 'bio.mrp.production.schedule.import.wizard'
    _description = 'MRP Production Shedule Import Wizard'

    @api.model
    def _default_warehouse_id(self):
        return self.env['stock.warehouse'].search([('company_id', '=', self.env.company.id)], limit=1)

    manufacturing_period = fields.Selection([
        ('month', 'Monthly'),
        ('week', 'Weekly')], string="Manufacturing Period",
        default='month', required=True,
        help="Default value for the time ranges in Master Production Schedule report.")
    warehouse_id = fields.Many2one('stock.warehouse', 'Production Warehouse',
        required=True, default=lambda self: self._default_warehouse_id())
    excel_file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    line_ids = fields.One2many(
        comodel_name='bio.mrp.production.schedule.lines.import.wizard',
        inverse_name='bio_mrp_production_schedule_wizard_id', string='Lines' )

    # Excel column configuration (0-indexed, A=0, B=1, C=2, etc.)
    header_row_number = fields.Integer(
        string='Header Row Number',
        default=1,
        required=True,
        help='Row number where column headers are located (1 = first row, 2 = second row, etc.)')
    default_code_column = fields.Integer(
        string='Product Code Column',
        default=1,
        required=True,
        help='Column number for product default_code (A=0, B=1, C=2, etc.)')
    product_name_column = fields.Integer(
        string='Product Name Column',
        default=4,
        required=True,
        help='Column number for product name (A=0, B=1, C=2, etc.)')
    first_date_column = fields.Integer(
        string='First Date Column',
        default=7,
        required=True,
        help='Column number where dates start (A=0, B=1, C=2, etc.)')

    @api.onchange('manufacturing_period')
    def _onchange_manufacturing_period(self):
        """Update column configuration based on manufacturing period"""
        # You can customize these defaults based on your Excel templates
        if self.manufacturing_period == 'month':
            # For monthly: typically B=code, E=name, H onwards=dates
            self.header_row_number = 1
            self.default_code_column = 1
            self.product_name_column = 4
            self.first_date_column = 7
        elif self.manufacturing_period == 'week':
            # For weekly: A=code, B=name, C onwards=dates
            self.header_row_number = 1
            self.default_code_column = 0
            self.product_name_column = 1
            self.first_date_column = 2


    def action_open_wizard(self):
        return {
            'name': _('Open MRP Production Shedule Import Wizard'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'bio.mrp.production.schedule.import.wizard',
            'target': 'new',
            'context': {},
        }

    def action_upload(self):
        """Parse Excel file and create wizard lines"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload an Excel file.'))

        # Check which library to use based on file extension
        filename = self.filename or ''
        use_xlrd = filename.endswith('.xls') or not filename.endswith('.xlsx')

        if use_xlrd and not xlrd:
            raise UserError(_('Python library "xlrd" is not installed. Please install it: pip install xlrd'))
        if not use_xlrd and not openpyxl:
            raise UserError(_('Python library "openpyxl" is not installed. Please install it: pip install openpyxl'))

        # Decode the file
        try:
            file_data = base64.b64decode(self.excel_file)
        except Exception as e:
            raise UserError(_('Error reading file: %s') % str(e))

        # Clear existing lines
        self.line_ids.unlink()

        if use_xlrd and xlrd:
            # Use xlrd for .xls files
            date_columns, lines_to_create = self._parse_excel_xlrd(file_data)
        else:
            # Use openpyxl for .xlsx files
            date_columns, lines_to_create = self._parse_excel_openpyxl(file_data)

        # Create lines
        if lines_to_create:
            self.env['bio.mrp.production.schedule.lines.import.wizard'].create(lines_to_create)

        # Return view with lines
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'bio.mrp.production.schedule.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _parse_excel_xlrd(self, file_data):
        """Parse Excel file using xlrd library (.xls format)"""
        try:
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
        except Exception as e:
            raise UserError(_('Error parsing Excel file with xlrd: %s') % str(e))

        # Convert header row number from 1-based to 0-based index
        header_row_idx = self.header_row_number - 1

        # Parse header row (dates)
        if sheet.nrows < header_row_idx + 1:
            raise UserError(_('Excel file does not have enough rows. Header row %d not found.') % self.header_row_number)

        header_row = sheet.row_values(header_row_idx)

        # Extract dates from header starting from configured column
        date_columns = []
        for col_idx in range(self.first_date_column, len(header_row)):
            cell_value = header_row[col_idx]
            if cell_value:
                try:
                    forecast_date = None
                    # Check if it's a date number (xlrd stores dates as float)
                    if isinstance(cell_value, float):
                        # Convert Excel date number to Python date
                        date_tuple = xldate.xldate_as_tuple(cell_value, workbook.datemode)
                        forecast_date = datetime(*date_tuple).date()
                    elif isinstance(cell_value, str):
                        # Try to parse date from DD.MM.YYYY format
                        forecast_date = datetime.strptime(cell_value.strip(), '%d.%m.%Y').date()

                    if forecast_date:
                        date_columns.append((col_idx, forecast_date))
                except (ValueError, xlrd.xldate.XLDateError):
                    # Skip invalid dates
                    continue

        if not date_columns:
            raise UserError(_('No valid dates found in header row %d starting from column %d. Expected format: DD.MM.YYYY (e.g., 01.12.2025)') % (self.header_row_number, self.first_date_column))

        # Parse data rows (starting from row after header)
        lines_to_create = []
        for row_idx in range(header_row_idx + 1, sheet.nrows):
            row = sheet.row_values(row_idx)

            # Get values from configured columns
            default_code = row[self.default_code_column] if self.default_code_column < len(row) else None
            product_name = row[self.product_name_column] if self.product_name_column < len(row) else ''

            if not default_code:
                continue  # Skip empty rows

            # Convert default_code to string
            if isinstance(default_code, float):
                default_code = str(int(default_code))
            else:
                default_code = str(default_code).strip()

            # Find product by default_code
            product = self.env['product.product'].search([('default_code', '=', default_code)], limit=1)

            # Create line for each date column
            for col_idx, forecast_date in date_columns:
                qty = row[col_idx] if col_idx < len(row) else 0

                # Skip if quantity is 0 or empty
                if not qty:
                    continue

                try:
                    forecast_qty = float(qty)
                except (ValueError, TypeError):
                    forecast_qty = 0.0

                if forecast_qty <= 0:
                    continue

                line_vals = {
                    'bio_mrp_production_schedule_wizard_id': self.id,
                    'default_code': default_code,
                    'product_name': str(product_name).strip() if product_name else '',
                    'forecast_date': forecast_date,
                    'forecast_qty': forecast_qty,
                }

                if product:
                    line_vals.update({
                        'product_id': product.id,
                        'state': 'found',
                        'message': _('Product found'),
                    })
                else:
                    line_vals.update({
                        'state': 'not_found',
                        'message': _('Product with code "%s" not found in database') % default_code,
                    })

                lines_to_create.append(line_vals)

        return date_columns, lines_to_create

    def _parse_excel_openpyxl(self, file_data):
        """Parse Excel file using openpyxl library (.xlsx format)"""
        try:
            file_obj = io.BytesIO(file_data)
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_('Error parsing Excel file with openpyxl: %s') % str(e))

        # Parse header row (dates) - header_row_number is 1-based, so use it directly with iter_rows
        header_row = list(sheet.iter_rows(min_row=self.header_row_number, max_row=self.header_row_number, values_only=True))[0]

        # Extract dates from header starting from configured column
        date_columns = []
        for col_idx, cell_value in enumerate(header_row[self.first_date_column:], start=self.first_date_column):
            if cell_value:
                try:
                    # Try to parse date from DD.MM.YYYY format
                    if isinstance(cell_value, str):
                        forecast_date = datetime.strptime(cell_value.strip(), '%d.%m.%Y').date()
                    elif isinstance(cell_value, datetime):
                        forecast_date = cell_value.date()
                    else:
                        continue
                    date_columns.append((col_idx, forecast_date))
                except ValueError:
                    # Skip invalid dates
                    continue

        if not date_columns:
            raise UserError(_('No valid dates found in header row %d starting from column %d. Expected format: DD.MM.YYYY (e.g., 01.12.2025)') % (self.header_row_number, self.first_date_column))

        # Parse data rows (starting from row after header)
        lines_to_create = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=self.header_row_number + 1, values_only=True), start=self.header_row_number + 1):
            # Get values from configured columns
            default_code = row[self.default_code_column] if self.default_code_column < len(row) else None
            product_name = row[self.product_name_column] if self.product_name_column < len(row) else ''

            if not default_code:
                continue  # Skip empty rows

            # Find product by default_code
            product = self.env['product.product'].search([('default_code', '=', str(default_code))], limit=1)

            # Create line for each date column
            for col_idx, forecast_date in date_columns:
                qty = row[col_idx] if col_idx < len(row) else 0

                # Skip if quantity is 0 or empty
                if not qty:
                    continue

                try:
                    forecast_qty = float(qty)
                except (ValueError, TypeError):
                    forecast_qty = 0.0

                if forecast_qty <= 0:
                    continue

                line_vals = {
                    'bio_mrp_production_schedule_wizard_id': self.id,
                    'default_code': str(default_code),
                    'product_name': str(product_name) if product_name else '',
                    'forecast_date': forecast_date,
                    'forecast_qty': forecast_qty,
                }

                if product:
                    line_vals.update({
                        'product_id': product.id,
                        'state': 'found',
                        'message': _('Product found'),
                    })
                else:
                    line_vals.update({
                        'state': 'not_found',
                        'message': _('Product with code "%s" not found in database') % default_code,
                    })

                lines_to_create.append(line_vals)

        return date_columns, lines_to_create

    def action_import(self):
        """Import validated lines to mrp.production.schedule"""
        self.ensure_one()

        lines_to_import = self.line_ids.filtered(lambda l: l.state == 'found')

        if not lines_to_import:
            raise UserError(_('No valid lines to import. Please upload and validate Excel file first.'))

        # TODO: Implement import to mrp.production.schedule
        # Group by product and create/update production schedule records

        imported_count = len(lines_to_import)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success'),
                'message': _('%d lines imported successfully. Import logic will be fully implemented soon.') % imported_count,
                'type': 'success',
                'sticky': False,
            }
        }


class MrpProductionSheduleLinessImportWizard(models.TransientModel):
    _name = 'bio.mrp.production.schedule.lines.import.wizard'
    _description = 'MRP Production Shedule Lines Import Wizard'
    _order = 'default_code, forecast_date'

    bio_mrp_production_schedule_wizard_id = fields.Many2one(
        comodel_name='bio.mrp.production.schedule.import.wizard',
        string='Wizard',
        required=True,
        ondelete='cascade')

    # Product information
    product_id = fields.Many2one(
        'product.product',
        string='Product',
        help='Product found by default_code')
    default_code = fields.Char(
        string='Product Code',
        required=True,
        help='Product code from Excel file (Column A)')
    product_name = fields.Char(
        string='Product Name',
        help='Product name from Excel file (Column B)')

    # Forecast data
    forecast_date = fields.Date(
        string='Forecast Date',
        required=True,
        help='Date from Excel header (DD.MM.YYYY format)')
    forecast_qty = fields.Float(
        string='Forecast Quantity',
        digits='Product Unit of Measure',
        required=True,
        help='Quantity from Excel cell')

    # Status
    state = fields.Selection([
        ('draft', 'Draft'),
        ('found', 'Product Found'),
        ('not_found', 'Product Not Found'),
        ('imported', 'Imported'),
        ('error', 'Error'),
    ], string='Status', default='draft', required=True)

    message = fields.Text(
        string='Message',
        help='Status message or error description')

